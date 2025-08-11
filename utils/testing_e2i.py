import openpyxl
from PIL import Image, ImageDraw, ImageFont
import os

class DirectExcelImageConverter:
    def __init__(self, cell_width=150, cell_height=40, font_size=14):  # Larger cells for Chinese
        self.cell_width = cell_width
        self.cell_height = cell_height
        self.font_size = font_size
    
    def _load_chinese_font(self, size):
        """Load a font that supports Chinese characters"""
        chinese_fonts = [
            # Windows
            "C:/Windows/Fonts/simsun.ttc",
            "C:/Windows/Fonts/simhei.ttf", 
            "C:/Windows/Fonts/msyh.ttc",
            
            # Linux (Docker)
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/arphic/uming.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",  # Fallback
        ]
        
        for font_path in chinese_fonts:
            try:
                if os.path.exists(font_path):
                    print(f"✅ Using font: {font_path}")
                    return ImageFont.truetype(font_path, size)
            except Exception as e:
                print(f"⚠️ Failed to load {font_path}: {e}")
                continue
        
        print("⚠️ Using default font - Chinese characters may not display correctly")
        return ImageFont.load_default()
    
    def _fit_text_in_cell(self, draw, text, font, cell_width):
        """Adjust text to fit in cell width"""
        text_width = draw.textlength(text, font=font)
        
        if text_width <= cell_width - 10:  # 10px padding
            return text, font
        
        # If text is too long, try smaller font
        smaller_font_size = max(8, self.font_size - 2)
        smaller_font = self._load_chinese_font(smaller_font_size)
        
        smaller_text_width = draw.textlength(text, font=smaller_font)
        if smaller_text_width <= cell_width - 10:
            return text, smaller_font
        
        # If still too long, truncate text
        max_chars = int(len(text) * (cell_width - 20) / text_width)
        truncated_text = text[:max_chars] + "..."
        return truncated_text, font
    
    def excel_to_image(self, excel_path: str, output_path: str, 
                      sheet_name: str = None) -> bool:
        """Convert Excel directly to image using openpyxl + PIL with Chinese support"""
        try:
            print(f"🚀 Converting {excel_path}")
            
            # Load workbook
            wb = openpyxl.load_workbook(excel_path, data_only=True)  # data_only=True for calculated values
            ws = wb[sheet_name] if sheet_name else wb.active
            
            # Get actual data range (skip empty rows/cols)
            max_row = ws.max_row
            max_col = ws.max_column
            
            print(f"📊 Processing {max_row} rows × {max_col} columns")
            
            # Calculate image dimensions
            img_width = max_col * self.cell_width
            img_height = max_row * self.cell_height
            
            # Create image with white background
            img = Image.new('RGB', (img_width, img_height), 'white')
            draw = ImageDraw.Draw(img)
            
            # Load Chinese-compatible font
            font = self._load_chinese_font(self.font_size)
            
            # Draw cells
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # Calculate position
                    x = (col - 1) * self.cell_width
                    y = (row - 1) * self.cell_height
                    
                    # Draw cell border
                    draw.rectangle(
                        [x, y, x + self.cell_width, y + self.cell_height],
                        outline='#666666', width=1
                    )
                    
                    # Draw cell background (header row)
                    if row == 1:
                        draw.rectangle(
                            [x + 1, y + 1, x + self.cell_width - 1, y + self.cell_height - 1],
                            fill='#4CAF50'
                        )
                        text_color = 'white'
                    else:
                        # Alternate row colors
                        if row % 2 == 0:
                            draw.rectangle(
                                [x + 1, y + 1, x + self.cell_width - 1, y + self.cell_height - 1],
                                fill='#f8f9fa'
                            )
                        text_color = 'black'
                    
                    # Draw text
                    if cell.value is not None:
                        text = str(cell.value).strip()
                        if text:  # Only draw non-empty text
                            # Fit text in cell
                            fitted_text, text_font = self._fit_text_in_cell(draw, text, font, self.cell_width)
                            
                            # Calculate text position (center vertically, left-align horizontally)
                            text_x = x + 8  # Left padding
                            text_y = y + (self.cell_height - self.font_size) // 2
                            
                            # Draw text
                            draw.text((text_x, text_y), fitted_text, fill=text_color, font=text_font)
            
            # Save image
            img.save(output_path, 'PNG', optimize=True, quality=95)
            print(f"✅ Image saved: {output_path}")
            
            wb.close()
            return True
            
        except Exception as e:
            print(f"❌ Direct conversion failed: {e}")
            import traceback
            traceback.print_exc()
            return False

# Test with improved version
if __name__ == "__main__":
    print("🚀 Starting conversion with Chinese font support...")
    
    excel2image_converter = DirectExcelImageConverter(
        cell_width=150,  # Wider for Chinese text
        cell_height=40,  # Taller for better readability
        font_size=12     # Good size for Chinese characters
    )
    
    converted = excel2image_converter.excel_to_image(
        excel_path="D:\\asianInfo\\ExcelAssist\\燕云村case\\测试用例v1(2)\\测试用例v1\\2021年中讯村村民小组长名单.xlsx",
        output_path="D:\\asianInfo\\dataProcessor\\utils\\test_chinese.png"
    )
    
    if converted:
        print("✅ Image created successfully!")
    else:
        print("❌ Image creation failed!")
